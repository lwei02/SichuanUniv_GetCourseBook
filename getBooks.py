import time
import json, os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import requests
from bs4 import BeautifulSoup
import openpyxl
import openpyxl.styles
from colorama import init, Fore, Back, Style

debug = False

# Login Info
SCHOOL_ID = ''
PASSWORD = ''

# OCR API Info (百度云OCR，选填，不填则人工输入验证码)
OCR_API_KEY = ''
OCR_SECRET_KEY = ''

init(autoreset=True)

def sel_session_to_requests_jar(wd_session : webdriver.Chrome):
    jar = requests.cookies.RequestsCookieJar()
    sel_cookies = wd_session.get_cookies()
    for cookie in sel_cookies:
        jar.set(cookie['name'], cookie['value'], domain=cookie['domain'], path=cookie['path'])
    return jar

if OCR_API_KEY and OCR_SECRET_KEY and not (OCR_API_KEY == '' or OCR_SECRET_KEY == ''):
    r = requests.get('https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + OCR_API_KEY + '&client_secret=' + OCR_SECRET_KEY)
    if r:
        access_token = r.json()['access_token']
        print(Fore.CYAN + ' ● ' + Fore.RESET + '获得百度云OCR访问令牌：' + access_token)
    else:
        print(Fore.RED + ' ● ' + Fore.RESET + '百度云OCR访问令牌获取失败。请检查OCR API Key和Secret Key是否正确。')
        exit()
else:
    print('未提供百度云OCR认证信息。将手动输入验证码。')
    captcha = ''
    access_token = None

service = Service('./lib/chromedriver.exe')
wco = webdriver.ChromeOptions()
wco.add_argument('log-level=3')
wco.add_experimental_option('excludeSwitches', ['enable-logging'])
wco.add_argument('--headless')
wco.add_argument('--disable-gpu')

wd = webdriver.Chrome(service=service, options=wco)
wd.get('http://id.scu.edu.cn/frontend/login')

while not (wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[1]/div/div/div[2]/div/input').is_displayed()):
    print('\r' + Fore.YELLOW + ' ● ' + Fore.RESET + '页面加载未完成，稍候……', end='')
    time.sleep(1)

print('\r' + Fore.GREEN + ' ● ' + Fore.RESET + '页面加载完成。执行操作。', end='')
id_input = wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[1]/div/div/div[2]/div/input')
id_input.send_keys(SCHOOL_ID)
pwd_input = wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[2]/div/div/div[2]/div/input')
pwd_input.send_keys(PASSWORD)
captcha_pic = wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[3]/div/div/img')
captcha_pic_src = captcha_pic.get_attribute('src')
captcha_pic.screenshot('./captcha.png')

if access_token != None:
    r = requests.post("https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token=" + access_token, data={'image': captcha_pic_src, 'language_type': 'ENG'}, headers={'Content-Type': 'application/x-www-form-urlencoded'})
    if r:
        captcha = ''
        for word in r.json()['words_result']:
            captcha += word['words']
        print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '验证码：' + captcha, end='')
    else:
        print('\r\n' + Fore.RED + ' ● ' + Fore.RESET + '百度云OCR识别失败。回退到手动输入验证码。', end='')
        captcha = ''

if captcha == '':
    os.startfile('captcha.png')
    captcha = input('\r\n' + Fore.MAGENTA + ' ● ' + Fore.RESET + '请输入验证码：')

captcha_input = wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[3]/div/div/div/div/input')
captcha_input.send_keys(captcha)
submit_btn = wd.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div/form/div[4]/div/button')
submit_btn.click()

print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '暂停操作3秒，等待数据传输完成……', end='')
time.sleep(3)

if wd.current_url == 'https://id.scu.edu.cn/frontend/login#/login':
    print('\r\n' + Fore.RED + ' ● ' + Fore.RESET + '登录失败，请检查验证信息或重试以排除验证码错误。', end='')
    exit()

print('\r\n' + Fore.GREEN + ' ● ' + Fore.RESET + '登录成功。\r\n\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '获取SSO信息……', end='')
os.remove('./captcha.png')
wd.get('https://id.scu.edu.cn/api/bff/v1.2/enduser/portal/sso/app_list')
sso_json = wd.find_element(By.XPATH, '/html/body/pre').text
sso_dict = json.loads(sso_json)
trm_f = 0
for app in sso_dict['data']['authorizationApplications']:
    if app['applicationId'] == 'scdxplugin_jwt23': #学生教务系统
        stu_link = app['startUrl']
        trm_f += 1
    if app['applicationId'] == 'scdxplugin_jwt2': #教师教务系统
        tea_link = app['startUrl']
        trm_f += 1
    if trm_f == 2:
        break
if trm_f != 2:
    print('\r\n' + Fore.YELLOW + ' ● ' + Fore.RESET + '获取SSO信息失败，将尝试使用预定义的链接。', end='')
    stu_link = 'https://id.scu.edu.cn/api/bff/v1.2/enduser/portal/sso/go_aca9e3773e3acaf330fa71729910e177cb3LCPYIaO9'
    tea_link = 'https://id.scu.edu.cn/api/bff/v1.2/enduser/portal/sso/go_784603c3ca458ef0b105f2cb44be944cGzZZ2ydWIbT'
else:
    print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '教师：' + stu_link, end='')
    print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '学生：' + tea_link, end='')
print('\r\n\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '正在选择学生系统以获得课程信息……', end='')

wd.get(stu_link)
print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '暂停操作3秒，等待数据传输完成……', end='')
time.sleep(3)
wd.get('http://zhjw.scu.edu.cn/student/courseSelect/calendarSemesterCurriculum/index')
while not wd.find_element(By.XPATH, '//*[@id="planCode"]').is_displayed():
    print('\r' + Fore.YELLOW + ' ● ' + Fore.RESET + '页面加载未完成，稍候……', end='')
    time.sleep(1)
print('\r\n' + Fore.GREEN + ' ● ' + Fore.RESET + '页面加载完成。执行操作。', end='')
plan_code_selector = wd.find_element(By.XPATH, '//*[@id="planCode"]')
# Get the plan code options' text and value
plan_code_options = plan_code_selector.find_elements(By.TAG_NAME, 'option')
plan_code_options_list = []
for option in plan_code_options:
    plan_code_options_list.append({
        'label': option.text,
        'value': option.get_attribute('value')
    })

stu_api_script = wd.find_element(By.XPATH, '/html/head/script[19]').get_attribute('innerHTML')
stu_api_key = stu_api_script.split('	url: "/student/courseSelect/thisSemesterCurriculum/')[1].split('/ajaxStudentSchedule/past/callback",')[0]
print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '课表查询API Key：' + stu_api_key, end='')

print('\r\n\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '*** USER SELECT ***')
print('' + Fore.CYAN + ' ● ' + Fore.RESET + '请选择学年学期：')
for item in plan_code_options_list:
    print(Fore.CYAN + ' ● ' + Fore.RESET + str(plan_code_options_list.index(item) + 1) + '. ' + item['label'])
print(Fore.CYAN + ' ● ' + Fore.RESET + '*******************\r\n', end='')
sel1_id = input(Fore.MAGENTA + ' ● ' + Fore.RESET + '选择：')
sel1_id = int(sel1_id) - 1
if sel1_id < 0 or sel1_id >= len(plan_code_options_list):
    print('\r\n' + Fore.YELLOW + ' ● ' + Fore.RESET + '选择无效。将选择1号。', end='')
    sel1_id = 0

sreq = requests.Session()
sreq.cookies = sel_session_to_requests_jar(wd)
stu_r = sreq.post('http://zhjw.scu.edu.cn/student/courseSelect/thisSemesterCurriculum/' + stu_api_key + '/ajaxStudentSchedule/past/callback',
    data={
        'planCode': plan_code_options_list[sel1_id]['value']
    }
)

stu_json = stu_r.json()
select_list = []
for program in stu_json['dateList']:
    for course in program['selectCourseList']:
        select_list.append({
            'name': course['courseName'],
            'teachers': [
                {
                    'name': tea.split(',')[1].split(' ')[0],
                    'id': tea.split(',')[0]
                } for tea in course['ywdgFlag'].split('|')
            ],
            'id': course['id']['coureNumber'], #此处为教务系统API拼写错误
            'seq': course['id']['coureSequenceNumber'],
            'program': program['programPlanName'],
            'program_code': program['programPlanCode'],
        })

print('\r\n\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '正在选择教师系统以查询教材信息……', end='')
wd.get(tea_link)
print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '暂停操作3秒，等待数据传输完成……', end='')
time.sleep(3)
sreq.cookies = sel_session_to_requests_jar(wd)
textbook_list = []
for course in select_list:
    print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '正在查询《' + course['name'] + '_' + course['seq'] +'》的教材信息……', end='')
    r = sreq.get('http://zhjwjs.scu.edu.cn/teacher/comprehensiveQuery/search/textbookSpecified/show?jsh=&kxh=' + course['seq'] + '&kch=' + course['id'] + '&zxjxjhh=' + plan_code_options_list[sel1_id]['value'])
    soup = BeautifulSoup(r.text, 'html.parser')
    table = soup.find('table', {'class': 'table table-striped table-bordered'})
    if table is None:
        print('\r' + Fore.RED + ' ● ' + Fore.RESET + '对《' + course['name'] + '_' + course['seq'] +'》的教材查询出错。            ', end='')
        continue
    if len(table.find_all('tr')) == 1:
        print('\r' + Fore.YELLOW + ' ● ' + Fore.RESET + '找到《' + course['name'] + '_' + course['seq'] +'》课程的0本教材。            ', end='')
        continue
    print('\r' + Fore.GREEN + ' ● ' + Fore.RESET + '找到《' + course['name'] + '_' + course['seq'] +'》课程的' + str(len(table.find_all('tr')) - 1) + '本教材。            ', end='')
    for row in table.find_all('tr')[1:]:
        textbook_list.append({
            'book_id': row.find_all('td')[0].text.strip(),
            'name': row.find_all('td')[1].text.strip(),
            'isbn': row.find_all('td')[2].text.strip(),
            'edition': row.find_all('td')[3].text.strip(),
            'printed': row.find_all('td')[4].text.strip(),
            'usage': row.find_all('td')[5].text.strip(),
            'as_for': row.find_all('td')[6].text.strip(),
            'needed': '是' if row.find_all('td')[7].text.strip() == '否' else '否',
            'course': course['name'],
            'course_id': course['id'],
            'course_seq': course['seq'],
            'program': course['program'],
            'program_code': course['program_code'],
            'teacher': ','.join([ tea['name']+"("+tea['id']+")" for tea in course['teachers']]) if len(course['teachers']) > 0 else '未知'
        })
print('\r\n\r\n' + Fore.GREEN + ' ● ' + Fore.RESET + '查询完毕，共找到' + str(len(textbook_list)) + '本教材。', end='')
if debug:
    print('\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '当前为调试状态，等待10秒后退出。', end='')
    time.sleep(10)
wd.quit()
print('\r\n\r\n' + Fore.CYAN + ' ● ' + Fore.RESET + '正在保存到文件……', end='')
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['课程号', '课程名称', '课序号', '教材编号', '教材名称', 'ISBN', '版本', '印次', '用途', '与课程关系', '是否必备', '教师', '所属教学计划', '教学计划代码'])
for book in textbook_list:
    ws.append([book['course_id'], book['course'], book['course_seq'], book['book_id'], book['name'], book['isbn'], book['edition'], book['printed'], book['usage'], book['as_for'], book['needed'], book['teacher'], book['program'], book['program_code']])
for cell in ws[1]:
    cell.font = openpyxl.styles.Font(name='FZHei-B01')
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = openpyxl.styles.Font(name='FZShuSong-Z01')
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 8
ws.column_dimensions['L'].width = 30
ws.column_dimensions['M'].width = 20
ws.column_dimensions['N'].width = 10

def save_file():
    global wb
    try:
        wb.save('textbook_' + plan_code_options_list[sel1_id]['value'] + '.xlsx')
    except PermissionError:
        print('\r\n' + Fore.RED + ' ● ' + Fore.RESET + '保存失败，文件被占用或权限不足。', end='')
        retry = input('\r\n' + Fore.MAGENTA + ' ● ' + Fore.RESET + '是否重试？(Y/N) ')
        if retry.lower() == 'y':
            save_file()
        else:
            exit()

save_file()
print('\r\n' + Fore.GREEN + ' ● ' + Fore.RESET + '保存完毕。', end='')



exit()

